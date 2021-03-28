import os
import pandas as pd
from pandas import ExcelWriter
from pandas.core.common import flatten

__all__ = ['get_page_tables','get_changes']

TIME_MODIFY = 'FFFF0000'
REFTABLE_XLSX = 'reftable.xlsx'

def get_page_tables_0(search_key, page):
    tables = pd.read_html(page)  # 'http://10.2.112.101:9001/ref/') #, match='Nominal GDP')
    with ExcelWriter('reftable.xlsx') as writer:  # ,mode='a'
        # https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.to_excel.html?highlight=to_excel#pandas.DataFrame.to_excel
        tables[1].to_excel(writer, sheet_name='sheet_1')
        tables[2].to_excel(writer, sheet_name='sheet_2')
        tables[3].to_excel(writer, sheet_name='sheet_3')
        tables[4].to_excel(writer, sheet_name='sheet_4')
        # https://pandas.pydata.org/docs/reference/api/pandas.read_excel.html


def get_page_tables(search_key, page):
    tables = pd.read_html(page)
    create_if_not_exist()
    for table in tables:
        if all(elem in list(flatten(list(table.columns.values))) for elem in ['Dia', 'Entrada', 'Horas Trabalhadas']):
            with pd.ExcelWriter(REFTABLE_XLSX, mode='a') as writer:
                # https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.to_excel.html?highlight=to_excel#pandas.DataFrame.to_excel
                table.to_excel(writer, sheet_name=''.join(e for e in search_key if e.isalnum()))
                writer.save()
                # https://pandas.pydata.org/docs/reference/api/pandas.read_excel.html
                print('Got table ' + search_key)
                print(table.head(4))
            return

    print('<< Table not found for ' + search_key + ' >>')


def create_if_not_exist(xlsx=REFTABLE_XLSX):
    if not os.path.exists(xlsx):
        w1 = pd.ExcelWriter(xlsx)
        pd.DataFrame({'Registro': []}).to_excel(w1, index=False)
        w1.save()
        w1.close()


def get_changes(mmyyyy='022014'):
    mmyyyy = mmyyyy.replace('/','')
    from openpyxl import load_workbook
    wb = load_workbook("C:\\Users\\vinicius.veo\\Downloads\\SeleniumLibrary\\atest\\" + REFTABLE_XLSX)
    print(wb.sheetnames)  # ['Sheet1', '022014', '032014', '042014', '052014']

    sheet = wb[mmyyyy]  # wb.active
    # print(sheet.max_column)
    # aaa = sheet.cell(1, 2).font.color
    res = []
    rows = sheet.rows
    for cells in rows:
        for cell in cells:
            font_color = cell.font.color
            cell_value = cell.value
            if font_color is None:
                font_color = ''
            else:
                font_color = font_color.value
            if cell_value is None:
                cell_value = ''
            value_n_color = str(cell_value) + ' - color: ' + str(font_color)
            coords = ' coords: [' + str(cell.row) + ',' + str(cell.column) + ']'
            week_day = sheet.cell(cell.row, 2).value
            record_type = sheet.cell(1, cell.column).value
            if font_color == TIME_MODIFY:
                print(value_n_color + coords + ' [' + week_day + '] ' + record_type)  # +str(rowCounter)+','+str(colCounter)+']') # +rows.index(cells)+','+cells.index(cell)+']') #
                xpath = 'xpath://*[@id="formConteudo:j_id267:opRelatorio:' + str(int(week_day.split('-')[0].strip()) - 1) + ':j_id305:0:j_id307"]'
                res.append([cell_value, font_color, week_day, record_type, xpath])

    return res


if __name__ == '__main__':
    get_changes()
